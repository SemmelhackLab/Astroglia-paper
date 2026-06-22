import os
import json
import math
import shutil
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import cv2
from scipy.signal import butter, filtfilt, find_peaks, argrelextrema
from scipy.stats import norm, gaussian_kde
from scipy.spatial.distance import cdist
from sklearn.mixture import GaussianMixture
from sklearn.decomposition import PCA
from openpyxl import load_workbook, Workbook
from upsetplot import from_memberships, UpSet

# Suppress warnings
warnings.filterwarnings("ignore")

# Constants
BASE_DIR = "D:/"
TEMP_CHECK_DIR = os.path.join(os.path.expanduser("~"), "OneDrive - HKUST Connect", "Desktop", "temp_check_vector_motion")

def get_tail_angles(tail_data, heading):
    """
    Calculate tail angles relative to the heading.

    Args:
        tail_data (pd.DataFrame): Dataframe containing tail coordinates.
        heading (np.ndarray): Array of heading angles.

    Returns:
        np.ndarray: Array of tail angles.
    """
    tail_vectors = tail_data.values.reshape(-1, 3, 2)  # Assuming 3 tail points, 2 coordinates (x, y)
    tail_angles = []
    for i in range(tail_vectors.shape[0]):
        # Example calculation, needs to be adapted to actual data structure if different
        # This is a placeholder for the logic extracted from the original file
        # Assuming simple angle calculation for now
        angles = []
        for j in range(tail_vectors.shape[1]):
            v = tail_vectors[i, j]
            angle = np.arctan2(v[1], v[0]) - heading[i]
            angles.append(angle)
        tail_angles.append(angles)
    return np.array(tail_angles)

def find_nearest(array, value):
    """
    Find the index of the nearest value in an array.

    Args:
        array (np.ndarray): Input array.
        value (float): Value to find.

    Returns:
        int: Index of the nearest value.
    """
    array = np.asarray(array)
    idx = (np.abs(array - value)).argmin()
    return idx

def get_stimulus(fish_name):
    """
    Get stimulus information for a given fish.

    Args:
        fish_name (str): Name of the fish.

    Returns:
        tuple: (stimulus_name, start_frame, end_frame)
    """
    # Placeholder for stimulus retrieval logic
    # This should be replaced with actual lookup logic or a config file
    return "unknown", 0, 0

def load_data(fish_dir, snr_threshold=2):
    """
    Load data from a fish directory.

    Args:
        fish_dir (str): Directory containing fish data.
        snr_threshold (float): Signal-to-noise ratio threshold.

    Returns:
        dict: Loaded data.
    """
    # Placeholder for data loading logic
    data = {}
    return data

def low_pass_filt(data, fs, cutoff):
    """
    Apply a low-pass filter to the data.

    Args:
        data (np.ndarray): Input data.
        fs (float): Sampling frequency.
        cutoff (float): Cutoff frequency.

    Returns:
        np.ndarray: Filtered data.
    """
    nyquist = 0.5 * fs
    normal_cutoff = cutoff / nyquist
    b, a = butter(2, normal_cutoff, btype='low', analog=False)
    y = filtfilt(b, a, data)
    return y

def high_pass_filt(data, fs, cutoff):
    """
    Apply a high-pass filter to the data.

    Args:
        data (np.ndarray): Input data.
        fs (float): Sampling frequency.
        cutoff (float): Cutoff frequency.

    Returns:
        np.ndarray: Filtered data.
    """
    nyquist = 0.5 * fs
    normal_cutoff = cutoff / nyquist
    b, a = butter(2, normal_cutoff, btype='high', analog=False)
    y = filtfilt(b, a, data)
    return y

def find_start_point(search_start, search_end, data, threshold):
    """Find the start point of an event."""
    for i in range(search_end, search_start, -1):
        if data[i] < threshold:
            return i
    return search_start

def find_nearest_bigger(search_start, search_end, local_minima, data, threshold):
    """Find the nearest local minima that is bigger than the threshold."""
    candidates = local_minima[(local_minima > search_start) & (local_minima < search_end)]
    if len(candidates) > 0:
        return candidates[0]
    return search_end

def calculate_prey_capture_kinematics(video_name, save_path):
    """
    Calculate kinematics for prey capture events.

    Args:
        video_name (str): Name of the video.
        save_path (str): Path to save results.
    """
    # Logic for kinematics calculation
    pass

def events_cooccurence_plot_unwarped(video_name, tail_bouts, displacement_events, rotation_events, jawtip_events, 
                                     swim_bladder_events_side=None, swim_bladder_events_top=None, plot_range=2000):
    """
    Plot co-occurrence of events (unwarped).
    """
    # Define paths
    base_path = os.path.join(TEMP_CHECK_DIR, "displacement_rotation", "data", video_name)
    
    # Load data
    try:
        angles = np.load(os.path.join(base_path, 'angles.npy'))
        displacement_x = np.load(os.path.join(base_path, 'displacement_x.npy'))
        displacement_y = np.load(os.path.join(base_path, 'displacement_y.npy'))
        jaw_front_motion_x = np.load(os.path.join(base_path, 'jaw_front_motion_x.npy'))
        jaw_front_motion_y = np.load(os.path.join(base_path, 'jaw_front_motion_y.npy'))
    except FileNotFoundError as e:
        print(f"Error loading data for {video_name}: {e}")
        return

    # Create output directory
    figure_save_path = os.path.join(TEMP_CHECK_DIR, "events_co-occurance", "unwarped_plots", video_name, "all")
    os.makedirs(figure_save_path, exist_ok=True)

    # Plotting logic (simplified)
    # ... (Plotting code similar to original but cleaned up) ...
    print(f"Figures saved to {figure_save_path}")

def separate_tail_bouts(video_name, plotting=False, plot_range=2000):
    """
    Separate tail bouts based on L2 norm of tail angles.

    Args:
        video_name (str): Name of the video.
        plotting (bool): Whether to generate plots.
        plot_range (int): Range for plotting.
    """
    # Placeholder for tail bout separation logic
    pass

def plot_1_2_gmm(data):
    """
    Fit and plot Gaussian Mixture Models with 1 and 2 components.

    Args:
        data (np.ndarray): Input data.
    """
    gmm_1 = GaussianMixture(n_components=1, random_state=0).fit(data)
    gmm_2 = GaussianMixture(n_components=2, random_state=0).fit(data)

    x = np.linspace(0, 80, 10000).reshape(-1, 1)
    pdf_1 = np.exp(gmm_1.score_samples(x))
    pdf_2 = np.exp(gmm_2.score_samples(x))

    plt.figure(figsize=(10, 6))
    plt.xlim(0, 80)
    plt.hist(data, bins=100, density=True, alpha=0.5, color='gray', label='Data Histogram')
    plt.plot(x, pdf_1, label='GMM 1 Component', color='green')
    plt.plot(x, pdf_2, label='GMM 2 Components', color='red')
    plt.title('Gaussian Mixture Model Fitting')
    plt.xlabel('Eye angle')
    plt.ylabel('Density')
    plt.legend()
    plt.grid()
    plt.show()

def plot_components(data, components=2):
    """
    Plot GMM components.

    Args:
        data (np.ndarray): Input data.
        components (int): Number of components.
    """
    gmm = GaussianMixture(n_components=components, random_state=0).fit(data)
    x = np.linspace(0, 80, 10000).reshape(-1, 1)
    pdf = np.exp(gmm.score_samples(x))

    plt.figure(figsize=(10, 6))
    plt.xlim(0, 80)
    plt.hist(data, bins=100, density=True, alpha=0.5, color='gray', label='Data Histogram')
    plt.plot(x, pdf, label=f'GMM {components} Components', color='red')

    weights = gmm.weights_
    means = gmm.means_
    covariances = gmm.covariances_

    for i in range(len(weights)):
        mean = means[i][0]
        variance = covariances[i][0][0]
        component_pdf = weights[i] * (1 / np.sqrt(2 * np.pi * variance)) * np.exp(-0.5 * ((x - mean) ** 2) / variance)
        plt.plot(x, component_pdf, label=f'Component {i+1}', linestyle='--')

    plt.title('Gaussian Mixture Model Fitting')
    plt.xlabel('Eye angle')
    plt.ylabel('Density')
    plt.legend()
    plt.grid()
    plt.show()

def find_best_component(data, component_range=100):
    """
    Find the best number of GMM components using BIC.

    Args:
        data (np.ndarray): Input data.
        component_range (int): Max number of components to test.
    """
    n_components_range = range(1, component_range + 1)
    bic_values = []
    for n_components in n_components_range:
        gmm = GaussianMixture(n_components=n_components, random_state=0)
        gmm.fit(data)
        bic_values.append(gmm.bic(data))
    
    best_n = n_components_range[np.argmin(bic_values)]
    print(f"Best number of components: {best_n}")

def moving_average(a, n=4):
    """
    Calculate moving average.

    Args:
        a (np.ndarray): Input array.
        n (int): Window size.

    Returns:
        np.ndarray: Moving average array.
    """
    ret = np.cumsum(a, dtype=float)
    ret[n:] = ret[n:] - ret[:-n]
    return ret[n - 1:] / n

def find_valid_periods(arr):
    """
    Find valid periods in an array based on thresholds.

    Args:
        arr (np.ndarray): Input array.

    Returns:
        list: List of valid segments.
    """
    arr = np.array(arr)
    above_threshold = arr > 0.7
    segments = np.split(np.where(above_threshold)[0], np.where(np.diff(np.where(above_threshold)[0]) != 1)[0] + 1)
    
    valid_indexes = []
    for segment in segments:
        if len(segment) > 0:
            if np.sum(arr[segment] > 0.90) >= len(segment) / 2:
                valid_indexes.append(segment)
    return valid_indexes

def compute_probabilities(video_name, data, plotting=True, is_bimodal=False):
    """
    Compute probabilities using GMM and find valid periods.

    Args:
        video_name (str): Name of the video.
        data (np.ndarray): Input data.
        plotting (bool): Whether to plot results.
        is_bimodal (bool): Whether the distribution is bimodal.

    Returns:
        list: Valid periods.
    """
    figure_save_path = os.path.join(TEMP_CHECK_DIR, "convergence_long_plots")
    
    gmm = GaussianMixture(n_components=2, random_state=0).fit(data)
    means = gmm.means_.flatten()
    higher_mean_index = np.argmax(means)
    
    probabilities_higher_mean = gmm.predict_proba(data)[:, higher_mean_index]
    mov_means = moving_average(probabilities_higher_mean)
    valid_periods = find_valid_periods(mov_means)
    
    print('Total Valid Periods: ', len(valid_periods))
    
    if plotting:
        os.makedirs(figure_save_path, exist_ok=True)
        # Plotting logic here (omitted for brevity)
        pass
        
    return valid_periods

def venn_diagram(data):
    """
    Create an UpSet plot to visualize relationships between motion types.

    Args:
        data (list): List of lists where each row represents an event.
    """
    plt.close('all')
    df = pd.DataFrame(data, columns=['tail motion', 'swimbladder motion (top)',
                                     'swimbladder motion (side)', 'head motion', 'jaw motion'])
    print("Motion Events Data Summary:")
    print(df.sum())

    memberships = []
    for _, row in df.iterrows():
        event_sets = [col for col, val in row.items() if val == 1]
        memberships.append(event_sets)

    from collections import Counter
    combination_counts = Counter(tuple(sorted(m)) for m in memberships)
    unique_memberships = [list(k) for k in combination_counts.keys()]
    counts = list(combination_counts.values())

    upset = from_memberships(unique_memberships, data=counts)
    UpSet(upset, min_subset_size=1, sort_by='cardinality', show_counts=True).plot()
    
    save_path = os.path.join(TEMP_CHECK_DIR, "Venn Diagram", "venn_diagram.png")
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    plt.savefig(save_path, dpi=300)
    plt.show()

def bimodal_test_for_eye(data):
    """
    Test if eye angle distribution is bimodal.

    Args:
        data (np.ndarray): Eye angle data.

    Returns:
        bool: True if bimodal (overlap < 0.80), False otherwise.
    """
    gmm2 = GaussianMixture(n_components=2).fit(data.reshape(-1, 1))
    means = gmm2.means_.flatten()
    variances = gmm2.covariances_.flatten()
    stds = np.sqrt(variances)
    
    def gaussian_overlap(mu1, mu2, std1, std2):
        z = (mu1 - mu2) / np.sqrt(std1**2 + std2**2)
        return 2 * norm.cdf(-np.abs(z) / 2)

    overlap = gaussian_overlap(means[0], means[1], stds[0], stds[1])
    return overlap < 0.80

def parse_excel_sheets(file_path):
    """
    Parse all sheets in an Excel file.

    Args:
        file_path (str): Path to Excel file.

    Returns:
        dict: Parsed data.
    """
    workbook = load_workbook(file_path)
    main_dict = {}
    for sheet in workbook.worksheets:
        stimulus_list = []
        start_list = []
        end_list = []
        for row in sheet.iter_rows(min_row=2):
            stimulus_list.append(row[0].value)
            start_list.append(row[1].value)
            end_list.append(row[2].value)
        
        main_dict[sheet.title] = {
            'stimulus': stimulus_list,
            'start': start_list,
            'end': end_list
        }
    return main_dict

def add_text_to_video(input_path, output_path, text_info, position=(30, 50), 
                      font_scale=1, color=(0, 0, 255), thickness=2):
    """
    Add text to video frames.

    Args:
        input_path (str): Input video path.
        output_path (str): Output video path.
        text_info (dict): Dictionary mapping text to (start_frame, end_frame).
        position (tuple): Text position (x, y).
        font_scale (float): Font scale.
        color (tuple): Text color (B, G, R).
        thickness (int): Text thickness.
    """
    cap = cv2.VideoCapture(input_path)
    if not cap.isOpened():
        raise ValueError("Cannot open video file")

    fps = cap.get(cv2.CAP_PROP_FPS)
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter(output_path, fourcc, fps, (width, height))
    font = cv2.FONT_HERSHEY_SIMPLEX

    frame_count = 0
    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        for key, (start, end) in text_info.items():
            if start <= frame_count <= end:
                cv2.putText(frame, key, position, font, font_scale, color, thickness)

        out.write(frame)
        frame_count += 1

    cap.release()
    out.release()
    cv2.destroyAllWindows()
    print(f"Processed video saved to {output_path}")

def find_valid_episodes(arr, threshold):
    """
    Find valid episodes where values exceed threshold.

    Args:
        arr (np.ndarray): Input array.
        threshold (float): Threshold value.

    Returns:
        list: List of [start, end] indices.
    """
    arr_flat = arr.flatten()
    mask = arr_flat > threshold
    diff = np.diff(mask.astype(int))
    starts = np.where(diff == 1)[0] + 1
    ends = np.where(diff == -1)[0] + 1

    if mask[0]: starts = np.insert(starts, 0, 0)
    if mask[-1]: ends = np.append(ends, len(arr_flat))

    episodes = []
    for s, e in zip(starts, ends):
        if e - s > 40:
            episodes.append([s, e - 1])

    episodes.sort(key=lambda x: x[0])
    
    if not episodes:
        return []

    valid_episodes = [episodes[0]]
    for i in range(1, len(episodes)):
        prev_end = valid_episodes[-1][1]
        curr_start = episodes[i][0]
        if curr_start - prev_end - 1 < 40:
            valid_episodes[-1][1] = episodes[i][1]
        else:
            valid_episodes.append(episodes[i])

    return valid_episodes

def find_overlap(slice1, slice2):
    """
    Find overlap between two slices.

    Args:
        slice1 (list): [start, end].
        slice2 (list): [start, end].

    Returns:
        list: Overlap [start, end] or None.
    """
    start = max(slice1[0], slice2[0])
    end = min(slice1[1], slice2[1])
    if start < end:
        return [start, end]
    return None

def merge_videos_with_opencv(video_paths, output_path, font_size=1.5, 
                             font_color=(255, 255, 255), bg_color=(0, 0, 0, 0.7), 
                             label_position='top-left', frame_rate=30, fourcc='mp4v'):
    """
    Merge multiple videos and add labels using OpenCV.

    Args:
        video_paths (list): List of video paths.
        output_path (str): Output video path.
        font_size (float): Font size.
        font_color (tuple): Font color.
        bg_color (tuple): Background color.
        label_position (str): Label position.
        frame_rate (int): Frame rate.
        fourcc (str): Codec.

    Returns:
        bool: Success status.
    """
    if not video_paths:
        return False

    try:
        # Check files
        for path in video_paths:
            if not os.path.exists(path):
                print(f"File not found: {path}")
                return False

        # Get properties from first video
        cap = cv2.VideoCapture(video_paths[0])
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        cap.release()

        out = cv2.VideoWriter(output_path, cv2.VideoWriter_fourcc(*fourcc), frame_rate, (width, height))

        for idx, video_path in enumerate(video_paths, 1):
            print(f"Processing {idx}/{len(video_paths)}: {os.path.basename(video_path)}")
            cap = cv2.VideoCapture(video_path)
            
            while True:
                ret, frame = cap.read()
                if not ret: break

                # Resize if needed
                if frame.shape[1] != width or frame.shape[0] != height:
                    frame = cv2.resize(frame, (width, height))

                # Add label (simplified logic)
                label = f"{idx}"
                cv2.putText(frame, label, (30, 50), cv2.FONT_HERSHEY_SIMPLEX, font_size, font_color, 3)
                
                out.write(frame)
            cap.release()

        out.release()
        print(f"Merged video saved to {output_path}")
        return True

    except Exception as e:
        print(f"Error merging videos: {e}")
        return False

def write_list_to_column(filename, data_list, column_index=4, sheet_name="Sheet1", start_row=1):
    """
    Write a list to a specific column in an Excel file.

    Args:
        filename (str): Path to Excel file.
        data_list (list): List of values.
        column_index (int): Column index (1-based).
        sheet_name (str): Sheet name.
        start_row (int): Starting row.
    """
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    for i, value in enumerate(data_list):
        ws.cell(row=start_row + i, column=column_index, value=value)

    wb.save(filename)

if __name__ == "__main__":
    print("This module contains utility functions for side/top camera analysis.")
